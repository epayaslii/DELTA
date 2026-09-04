"""Generate docs/articles-by-pipeline-part.xlsx -- the paper -> pipeline-part map.

Source of truth: docs/implementation-plan.md. Regenerate after any change there:
    .venv/bin/python scripts/make_articles_xlsx.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs" / "articles-by-pipeline-part.xlsx"

HEADERS = ["Pipeline part", "Article", "Link", "ID / venue", "Authors",
           "Role", "What we take", "What we DON'T take / how we adapt", "Status"]

ROWS = [
    # part, article, link, id/venue, authors, role, take, adapt, status
    ("Base task + downstream interface", "DELTA", "",
     "Maté & Dimiccoli (IRI.pdf, internal)", "Maté, Dimiccoli", "Core V1",
     "Dense long-term anticipation from transcripts; Y* format; the decoder we feed",
     "Its ATBA-based TA (we replace that)", "interface fixed"),
    ("Stage 0 - semantic-guided sampling", "LGTTP",
     "https://aclanthology.org/2025.emnlp-main.451.pdf",
     "EMNLP 2025", "Y. Kumar", "Core V1",
     "Language decides where to spend VLM compute; dense tokens near transitions",
     "Its VideoLLM-QA task wrapper", "spec"),
    ("Stage A1 - VLM semantic evidence", "OVTAS",
     "https://arxiv.org/abs/2602.21406",
     "arXiv 2602.21406 (Feb 2026)", "(open-vocab zero-shot AS study)", "Core V1",
     "FAES: frozen VLM clip<->action cosine; 14-VLM comparison guides backbone choice",
     "Its unordered / open-vocab label-set assumption (we keep transcript order)",
     "built (delta.align.cost)"),
    ("Stage A2 - clips not frames", "HiERO",
     "https://openaccess.thecvf.com/content/ICCV2025/papers/Peirone_HiERO_Understanding_the_Hierarchy_of_Human_Behavior_Enhances_Reasoning_on_ICCV_2025_paper.pdf",
     "ICCV 2025", "Peirone, Pistilli, Averta", "Core V1",
     "Represent actions with temporal context; top-down activity hierarchy",
     "Its egocentric narration/Ego4D training pipeline", "spec (feature choice)"),
    ("Stage A3 - ordered monotonic decode", "HiERO-StepG",
     "https://arxiv.org/abs/2605.31227",
     "arXiv 2605.31227", "Zenotto, Peirone, Pistilli, Averta", "Core V1",
     "Ordered step descriptions -> strict-monotonic grounding, zero-shot; noise post-proc",
     "HiERO features / the full Ego4D pipeline", "built (delta.align.asot.decode)"),
    ("Stage A engine - OT solver", "ASOT",
     "https://arxiv.org/abs/2404.01518",
     "CVPR 2024", "Xu & Gould", "Core V1",
     "Fused Gromov-Wasserstein OT with a temporal prior (in WLTA wclot)",
     "Unsupervised cluster identities (we use transcript entries)",
     "built (delta.align.asot)"),
    ("Stage A engine - OT solver", "CLOT",
     "https://openaccess.thecvf.com/content/ICCV2025/papers/Bueno-Benito_CLOT_Closed_Loop_Optimal_Transport_for_Unsupervised_Action_Segmentation_ICCV_2025_paper.pdf",
     "ICCV 2025", "Bueno-Benito & Dimiccoli", "Core V1",
     "Closed-loop frame<->segment OT refinement (the group's wclot code)",
     "Unsupervised cluster identities", "built (delta.align.asot)"),
    ("Stage B1 - local boundary search", "OUR DESIGN", "", "-", "-",
     "Core V1 - NOVEL",
     "+-r window per coarse boundary; score 'left=A not B, right=B not A' + visual change; pick best + confidence",
     "-", "built (delta.align.refine)"),
    ("Stage B2 - PBCR contrastive refinement", "CVA (principle only)",
     "https://github.com/byeol3325/CVA_cvpr",
     "CVPR 2026 (cva_cvpr.pdf)", "Moon, Hyun, Lee, Heo", "Core V1 / V2",
     "The idea that a boundary deserves its own contrastive representation",
     "Published CBD anchors on GT-span boundaries + GT-relative negatives + GT matching; "
     "we rebuild anchors from CONFIDENT pseudo-boundaries -> our objective (PBCR)",
     "built (delta.align.cbd)"),
    ("Stage B - relational sharpening", "MASRA",
     "https://arxiv.org/abs/2605.03398",
     "arXiv 2605.03398 (ACM, 2026)",
     "Ran, Wei, Zhou, Qin, He, Ma, Zhou, Yang", "Core V1",
     "LRCA / ESTA train-time regularizers on local windows",
     "Its full DETR VTG backbone; GT spans; DAI/SGE/SORA (backbone-specific)",
     "built (delta.align.masra_torch)"),
    ("VLM backbone (all stages)", "VideoLLaMA3",
     "https://arxiv.org/abs/2501.13106",
     "arXiv 2501.13106 (2025)", "Zhang et al.", "Core V1",
     "SigLIP-based vision tower -> similarity s; Chat model -> captions/reasoning. "
     "Supervisor's call (2026-09-03), explicitly NOT InternVideo2",
     "-", "cluster (M-A4)"),
    # --- Later versions ----------------------------------------------------
    ("Stage A alternative (ablation)", "TASOT",
     "https://arxiv.org/abs/2602.24138",
     "arXiv 2602.24138",
     "Mohamed, Fazzari, Al-Naji, Alhadhrami, Hableel, Alkindi, Laptev, Stefanini",
     "Ablation",
     "Multimodal OT: fuse visual structure + VLM-caption semantics + temporal reg in one cost",
     "Surgical domain; generated descriptions as-is (we use transcript actions, impose order)",
     "cluster ablation"),
    ("Stage B3 - hard-case reasoning", "TOGA",
     "https://arxiv.org/abs/2506.09445",
     "arXiv 2506.09445 (Jun 2025)",
     "Gupta, Roy, Chellappa, Bastian, Velasquez, Jha", "V4",
     "Weakly-supervised temporal grounding: pseudo-labels + consistency, no GT timestamps; "
     "accepts prompts with [start,end] -> 'segments in the VLM input'",
     "Its VideoQA / one-query-one-span task; run only locally on low-confidence boundaries",
     "spec"),
    ("Stage C - iterative frame<->segment", "D-CLOT",
     "https://arxiv.org/abs/2608.05877",
     "arXiv 2608.05877 (TPAMI sub.)", "Bueno-Benito, Dimiccoli", "V3",
     "Prototype re-estimation + graph geometry; explicit focus on ambiguous transitions & short actions",
     "Full unsupervised clustering pipeline (we keep transcript-defined segments)",
     "spec"),
    ("Optional visual-structure features", "V-JEPA2",
     "https://arxiv.org/abs/2506.09985",
     "Meta 2025 (features via D-CLOT release)", "Assran et al. / D-CLOT", "Optional",
     "Pure-vision video embeddings as a structure term alongside VideoLLaMA3",
     "-", "cluster (optional)"),
    ("Dynamic structural prior tweak", "FIS-OT",
     "https://arxiv.org/abs/2608.29980",
     "arXiv 2608.29980 (github flying05/FIS-OT)",
     "Peng, Qin, Li, Yang, Wang", "Reference",
     "Replace ASOT's rigid temporal mask with temporal-mask + feature-affinity blend; "
     "triplet loss for local consistency independent of pseudo-labels",
     "Full unsupervised setup", "reference"),
    # --- Baselines (comparison, not method) -------------------------------
    ("Baseline - no-evidence floor", "naive-uniform", "", "-", "-", "Baseline",
     "Split each video into len(transcript) equal parts in order", "-",
     "measured: MoC 0.366"),
    ("Baseline - classifier-based TA", "ATBA",
     "https://arxiv.org/abs/2403.19225",
     "CVPR 2024 (iSEE-Laboratory/CVPR24_ATBA)", "Xu, Zheng", "Baseline",
     "'Alignment through segmentation' number, via WLTA --model_type atba", "-",
     "cluster"),
    ("Baseline - OT-based TA", "ASOT / CLOT",
     "https://arxiv.org/abs/2404.01518",
     "CVPR 2024 / ICCV 2025", "Xu & Gould / Bueno-Benito & Dimiccoli", "Baseline",
     "Via WLTA --model_type wclot", "-", "cluster"),
    ("Baseline - segmentation SOTA (cite only)", "HAL",
     "https://arxiv.org/abs/2602.24275",
     "arXiv 2602.24275 (CVPR 2026)", "Huang et al.", "Baseline",
     "= ATBA + a VAE regularizer; Breakfast MoF number only (never ran 50Salads)",
     "Dropped as a workstream (no VLM, segmentation-based)", "cite"),
    # --- Supporting citations --------------------------------------------
    ("Paradigm citation (CVPR stamp)", "MLLM4WTAL",
     "https://arxiv.org/abs/2411.08466",
     "CVPR 2025", "Zhang et al.", "Citation",
     "'MLLM guides a weakly-supervised model at training only'", "-", "cite"),
    ("Aligner mechanism reference", "StepFormer",
     "https://arxiv.org/abs/2304.13265", "CVPR 2023", "Dvornik et al.", "Citation",
     "Ordered step slots -> video, order-aware alignment loss, self-supervised", "-", "cite"),
    ("Aligner mechanism reference", "TAN (Temporal Alignment Networks)",
     "https://arxiv.org/abs/2204.02968", "CVPR 2022 Oral", "Han, Xie, Zisserman",
     "Citation", "Sentences -> long video, weak from noisy narration", "-", "cite"),
    ("Aligner mechanism reference", "Drop-DTW",
     "https://arxiv.org/abs/2108.11996", "NeurIPS 2021", "Dvornik et al.", "Citation",
     "Differentiable sequence<->video DTW that can drop outliers", "-", "cite"),
    ("Joint video-video align + segmentation", "VAOT / VASOT",
     "https://arxiv.org/abs/2503.16832", "arXiv 2503.16832",
     "Ali, Mahmood, Saeed, Konin, Zia, Tran", "Citation",
     "Fused-GW OT with structural priors; single model for alignment + segmentation",
     "Alignment = video<->video, not transcript<->video", "cite"),
]

FILLS = {
    "Core V1": "C6E0B4", "Core V1 - NOVEL": "FFE699", "Core V1 / V2": "C6E0B4",
    "V2": "BDD7EE", "V3": "BDD7EE", "V4": "BDD7EE",
    "Ablation": "D9D9D9", "Optional": "D9D9D9", "Reference": "D9D9D9",
    "Baseline": "F8CBAD", "Citation": "EDEDED",
}

# ---- one-glance summary sheet (for the supervisor) --------------------------
# Stage, What we do, Main source of inspiration, What we actually borrow, Links
SUMMARY_HEADERS = ["Stage", "What we do", "Main source of inspiration",
                   "What we actually borrow", "Links"]
SUMMARY = [
    ("Input", "Raw video + ordered transcript", "DELTA",
     "Same weak supervision: no timestamps / boundaries / durations",
     "Mate & Dimiccoli (internal)"),
    ("Semantic front-end",
     "Encode short video clips and transcript actions into a shared semantic space",
     "OVTAS",
     "Frozen VLM visual/text similarity idea",
     "https://arxiv.org/abs/2602.21406"),
    ("Stage A: coarse alignment",
     "Build the cost and force actions to appear in transcript order",
     "our existing monotonic DP + HiERO-StepG principle",
     "Strict monotonic ordered alignment",
     "https://arxiv.org/abs/2605.31227"),
    ("Stage B1: segment extent",
     "Find the confident temporal extent of each action around its coarse location",
     "HiERO-StepG",
     "Query-conditioned boundary / segment expansion",
     "https://arxiv.org/abs/2605.31227"),
    ("Stage B2: exact transition",
     "Search specifically for the best crossover point",
     "our formulation (+ CVA boundary-contrastive principle, MASRA LRCA)",
     "Explicit GT-free semantic boundary localization",
     "https://github.com/byeol3325/CVA_cvpr ; https://arxiv.org/abs/2605.03398"),
    ("Optional Stage C",
     "Improve frame / segment consistency if boundaries remain noisy",
     "CLOT / D-CLOT",
     "Closed-loop segment / frame / prototype refinement",
     "https://openaccess.thecvf.com/content/ICCV2025/papers/Bueno-Benito_CLOT_Closed_Loop_Optimal_Transport_for_Unsupervised_Action_Segmentation_ICCV_2025_paper.pdf ; https://arxiv.org/abs/2608.05877"),
    ("Optional Stage D",
     "Reason only about difficult boundaries",
     "TOGA / VideoLLaMA3",
     "Expensive local VLM reasoning, selectively",
     "https://arxiv.org/abs/2506.09445 ; https://arxiv.org/abs/2501.13106"),
    ("Output", "Convert refined boundaries into dense Y*", "DELTA",
     "Preserve the existing downstream interface", "Mate & Dimiccoli (internal)"),
]


def build():
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary (for Mariella)"
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="305496")
    for c, h in enumerate(SUMMARY_HEADERS, 1):
        cell = ws0.cell(1, c, h)
        cell.font, cell.fill = hfont, hfill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r, row in enumerate(SUMMARY, 2):
        for c, val in enumerate(row, 1):
            cell = ws0.cell(r, c, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([22, 40, 30, 36, 46], 1):
        ws0.column_dimensions[get_column_letter(i)].width = w
    ws0.freeze_panes = "A2"
    ws0.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{len(SUMMARY) + 1}"

    ws = wb.create_sheet("Full detail")

    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="305496")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    link_font = Font(color="0563C1", underline="single")
    for r, row in enumerate(ROWS, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[2]:
            lc = ws.cell(r, 3)
            lc.hyperlink = row[2]
            lc.font = link_font
        role = row[5]
        if role in FILLS:
            ws.cell(r, 6).fill = PatternFill("solid", fgColor=FILLS[role])

    widths = [30, 22, 42, 26, 26, 14, 44, 44, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    wb.save(OUT)
    print(f"wrote {OUT}  ({len(ROWS)} rows)")


if __name__ == "__main__":
    build()
